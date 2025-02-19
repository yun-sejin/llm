from openpyxl import Workbook

# Create a workbook and active worksheet
wb = Workbook()
ws = wb.active

# Add headers and sample rows
ws['A1'] = "Header1"
ws['B1'] = "Header2"
ws.append(["Row1 - Column1", "Row1 - Column2"])
ws.append(["Row2 - Column1", "Row2 - Column2"])

# Save the workbook to a file
output_file = "example.xlsx"
wb.save(output_file)
print(f"Excel file '{output_file}' generated successfully.")
